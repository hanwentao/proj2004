import pathlib

from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model

import openpyxl


class Writer:

    def __init__(self, template_path, output_path):
        self.book = openpyxl.load_workbook(template_path)
        self.sheet = self.book.active
        self.output_path = output_path

    def append(self, row):
        self.sheet.append(row)

    def close(self):
        self.book.save(self.output_path)


class WriterGroup:

    def __init__(self, template_path, summary_template_path, base_path):
        self.template_path = template_path
        self.base_path = pathlib.Path(base_path)
        self.base_path.mkdir(parents=True, exist_ok=True)
        self.last_department = None
        self.last_class = None
        self.all_writer = Writer(template_path, self.base_path / '2004级.xlsx')
        self.summary_writer = Writer(summary_template_path, self.base_path / '统计.xlsx')
        self.department_writer = None
        self.class_writer = None

    def check_new_group(self, department=None, class_=None):
        if department != self.last_department:
            if self.department_writer is not None:
                self.department_writer.close()
            if department is not None:
                print(department)
                self.department_path = self.base_path / department
                self.department_path.mkdir(parents=True, exist_ok=True)
                self.department_writer = Writer(self.template_path, self.department_path / f'{department}.xlsx')
        if class_ != self.last_class:
            if self.class_writer is not None:
                self.class_writer.close()
                self.summary_writer.append([
                    self.last_department,
                    self.last_class,
                    self.counter,
                    self.login_counter,
                    self.attend_counter,
                    self.card_counter,
                    self.email_counter,
                ])
            self.last_class = class_
            if class_ is not None:
                print(class_)
                self.class_writer = Writer(self.template_path, self.department_path / f'{class_}.xlsx')
                self.counter = 0
                self.login_counter = 0
                self.attend_counter = 0
                self.card_counter = 0
                self.email_counter = 0
        if department != self.last_department:
            self.last_department = department

    def append(self, row):
        department = row[0]
        class_ = row[1]
        self.check_new_group(department, class_)
        self.all_writer.append(row)
        self.department_writer.append(row)
        self.class_writer.append(row)
        self.counter += 1
        if row[4] == '是': self.login_counter += 1
        if row[6] == '参加': self.attend_counter += 1
        if row[7] == '是': self.card_counter += 1
        if row[8] != '未申请': self.email_counter += 1

    def close(self):
        self.check_new_group()
        self.all_writer.close()
        self.department_writer.close()
        self.class_writer.close()
        self.summary_writer.close()


class Command(BaseCommand):
    help = 'Export alumni list.'

    def add_arguments(self, parser):
        parser.add_argument('-t', '--template', required=True, help='The list template.')
        parser.add_argument('-T', '--summary-template', required=True, help='The summary template.')
        parser.add_argument('-s', '--source', required=True, help='The source list.')
        parser.add_argument('-O', '--output', required=True, help='The output path.')

    def handle(self, *args, **options):
        User = get_user_model()
        book = openpyxl.load_workbook(options['source'])
        sheet = book.active
        writer_group = WriterGroup(options['template'], options['summary_template'], options['output'])
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            department = row[1].value
            student_id = row[2].value
            name = row[3].value
            class_ = row[6].value
            user = User.objects.get(username=student_id)
            profile = user.profile
            extra = user.extra
            writer_group.append([
                department,
                class_,
                student_id,
                name,
                '是' if user.last_login else '否',
                f'{profile.completeness * 100:0.0f}%',
                '参加' if extra.attend else ('不参加' if extra.attend is False else '未填写'),
                '是' if extra.photo else '否',
                f'{extra.email_prefix}04@tsinghua.org.cn' if extra.email_prefix else '未申请',
            ])
            self._check_equal(profile.name, name, student_id)
            self._check_equal(profile.department_name, department, student_id)
            self._check_equal(profile.class_name, class_, student_id)
        writer_group.close()

    def _check_equal(self, left, right, id):
        if left != right:
            self.stdout.write(self.style.WARNING(f'{id}: {left} != {right}'))
