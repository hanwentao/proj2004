import os

from django.core.management.base import BaseCommand, CommandError
from django.conf import settings
from django.db import connection

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ...models import Class


class Command(BaseCommand):
    help = 'Export invitation in Excel.'

    def add_arguments(self, parser):
        parser.add_argument('-p', '--path', default='.', help='Specifies path of output files.')

    def handle(self, *args, **options):
        path = options['path']
        for class_ in Class.objects.all():
            class_name = class_.name
            department_name = class_.department.name
            xlsx_path = f'{path}/{department_name}/{class_name}.xlsx'
            print(f'Generating {xlsx_path}...')
            wb = Workbook()
            ws = wb.active
            ws.title = '名单'
            ws.append([
                '学号',
                '姓名',
                '院系',
                '班级',
                '邀请链接',
            ])
            ws.column_dimensions[get_column_letter(1)].width = 15
            ws.column_dimensions[get_column_letter(2)].width = 15
            ws.column_dimensions[get_column_letter(3)].width = 20
            ws.column_dimensions[get_column_letter(4)].width = 10
            ws.column_dimensions[get_column_letter(5)].width = 60
            for profile in class_.profile_set.all():
                ws.append([
                    profile.student_id,
                    profile.name,
                    department_name,
                    class_name,
                    profile.invitation_url,
                ])
            os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)
            wb.save(xlsx_path)
