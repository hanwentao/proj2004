from django.core.management.base import BaseCommand, CommandError
from django.conf import settings
from django.db import connection

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ...models import Profile


class Command(BaseCommand):
    help = 'Export invitation in Excel.'

    def add_arguments(self, parser):
        parser.add_argument('-p', '--prefix', help='Specifies prefix of output files.')
        parser.add_argument('-d', '--department', action='store_true', help='Exports by departments.')
        parser.add_argument('-c', '--class', action='store_true', help='Exports by classes.')

    def handle(self, *args, **options):
        prefix = options['prefix']
        by_department = options['department']
        by_class = options['class']
        if not by_department and not by_class:
            raise CommandError('Must specify either by deparment or by class')
        if by_department and by_class:
            raise CommandError('Cannot specify both by department and by class')
        with connection.cursor() as cursor:
            if by_department:
                cursor.execute('SELECT DISTINCT department FROM contacts_profile')
            else:
                cursor.execute('SELECT DISTINCT clazz FROM contacts_profile')
            groups = cursor.fetchall()
        groups = [g[0] for g in groups]
        prefix = '' if prefix is None else prefix
        for group in groups:
            name = '空' if not group else group
            if not prefix or prefix.endswith('/'):
                path = f'{prefix}{name}.xlsx'
            else:
                path = f'{prefix}-{name}.xlsx'
            print(f'Generating {path}...')
            wb = Workbook()
            ws = wb.active
            ws.title = '名单'
            ws.append([
                '学号',
                '姓名',
                '院系',
                '班级',
                '邀请链接',
            ])
            ws.column_dimensions[get_column_letter(1)].width = 15
            ws.column_dimensions[get_column_letter(2)].width = 15
            ws.column_dimensions[get_column_letter(3)].width = 20
            ws.column_dimensions[get_column_letter(4)].width = 10
            ws.column_dimensions[get_column_letter(5)].width = 60
            if by_department:
                alumni = Profile.objects.filter(department=group)
            else:
                alumni = Profile.objects.filter(clazz=group)
            for i, alumnus in enumerate(alumni):
                ws.append([
                    alumnus.student_id,
                    alumnus.name,
                    alumnus.department,
                    alumnus.clazz,
                    settings.BASE_URL + alumnus.get_absolute_url() + 'edit/?code=' + alumnus.verification_code,
                ])
            wb.save(path)
